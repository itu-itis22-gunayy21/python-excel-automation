"""
📊 Python Excel Automation — Sales Report Generator
Reads raw sales data from Excel, cleans it, analyzes it,
and produces a professional report with charts.
"""

import os
import sys
import datetime
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows


# ── colour palette ──────────────────────────────────────────────
DARK   = "1B2A4A"
ACCENT = "2E86DE"
GREEN  = "27AE60"
ORANGE = "F39C12"
WHITE  = "FFFFFF"
LIGHT  = "EBF5FB"

header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
header_fill = PatternFill("solid", fgColor=DARK)
accent_fill = PatternFill("solid", fgColor=LIGHT)
thin_border = Border(
    left=Side("thin"), right=Side("thin"),
    top=Side("thin"), bottom=Side("thin"),
)
money_fmt = '#,##0.00 ₺'


def style_header(ws, col_count):
    for col in range(1, col_count + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border


def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)


# ── STEP 1 — Load & Clean ──────────────────────────────────────
def load_and_clean(path: str) -> pd.DataFrame:
    print("📂 Loading data …")
    df = pd.read_excel(path)
    df.columns = df.columns.str.strip()

    # Normalise column names
    col_map = {}
    for c in df.columns:
        low = c.lower()
        if "order" in low and "id" in low:   col_map[c] = "OrderID"
        elif "date" in low:                  col_map[c] = "Date"
        elif "product" in low:               col_map[c] = "Product"
        elif "categ" in low:                 col_map[c] = "Category"
        elif "quant" in low:                 col_map[c] = "Quantity"
        elif "unit" in low:                  col_map[c] = "UnitPrice"
        elif "total" in low:                 col_map[c] = "Total"
        elif "city" in low:                  col_map[c] = "City"
        elif "sales" in low and "rep" in low:col_map[c] = "SalesRep"
        elif "payment" in low:               col_map[c] = "Payment"
        elif "customer" in low:              col_map[c] = "CustomerType"
        elif "discount" in low:              col_map[c] = "Discount"
    df.rename(columns=col_map, inplace=True)

    # Parse dates
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        df.dropna(subset=["Date"], inplace=True)
        df["Month"]   = df["Date"].dt.to_period("M")
        df["MonthStr"] = df["Date"].dt.strftime("%Y-%m")

    # Ensure numeric
    for c in ["Quantity", "UnitPrice", "Total", "Discount"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    # Remove negative / zero totals
    if "Total" in df.columns:
        df = df[df["Total"] > 0]

    print(f"   ✅ {len(df)} clean rows loaded.")
    return df


# ── STEP 2 — Analyse ───────────────────────────────────────────
def analyse(df: pd.DataFrame) -> dict:
    print("🔍 Analysing …")
    res = {}

    res["total_revenue"]  = df["Total"].sum()
    res["total_orders"]   = len(df)
    res["avg_order"]      = df["Total"].mean()
    res["median_order"]   = df["Total"].median()

    # Top 10 products by revenue
    res["top_products"] = (
        df.groupby("Product")["Total"]
        .sum().sort_values(ascending=False).head(10)
    )

    # Revenue by category
    res["category_rev"] = df.groupby("Category")["Total"].sum().sort_values(ascending=False)

    # Monthly revenue
    if "MonthStr" in df.columns:
        res["monthly_rev"] = df.groupby("MonthStr")["Total"].sum().sort_index()

    # Top sales reps
    if "SalesRep" in df.columns:
        res["top_reps"] = (
            df.groupby("SalesRep")["Total"]
            .sum().sort_values(ascending=False).head(5)
        )

    # City breakdown
    if "City" in df.columns:
        res["city_rev"] = df.groupby("City")["Total"].sum().sort_values(ascending=False)

    # Payment method split
    if "Payment" in df.columns:
        res["payment_split"] = df.groupby("Payment")["Total"].sum()

    # Customer type split
    if "CustomerType" in df.columns:
        res["customer_split"] = df.groupby("CustomerType")["Total"].sum()

    print("   ✅ Analysis complete.")
    return res


# ── STEP 3 — Charts (saved as PNG) ─────────────────────────────
def create_charts(res: dict, out_dir: str) -> dict:
    print("📊 Creating charts …")
    os.makedirs(out_dir, exist_ok=True)
    paths = {}
    palette = ["#2E86DE","#27AE60","#F39C12","#E74C3C","#8E44AD",
               "#1ABC9C","#D35400","#2C3E50","#C0392B","#16A085"]

    plt.rcParams.update({"font.family": "sans-serif", "font.size": 11})

    # 1 — Monthly Revenue
    if "monthly_rev" in res:
        fig, ax = plt.subplots(figsize=(10, 5))
        res["monthly_rev"].plot(kind="bar", color=palette[0], ax=ax, edgecolor="white")
        ax.set_title("Monthly Revenue", fontsize=14, fontweight="bold")
        ax.set_ylabel("Revenue (₺)")
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{x:,.0f}'))
        plt.xticks(rotation=45, ha="right")
        plt.tight_layout()
        p = os.path.join(out_dir, "monthly_revenue.png")
        fig.savefig(p, dpi=150)
        plt.close(fig)
        paths["monthly"] = p

    # 2 — Top Products
    fig, ax = plt.subplots(figsize=(10, 5))
    res["top_products"].plot(kind="barh", color=palette[1], ax=ax, edgecolor="white")
    ax.set_title("Top 10 Products by Revenue", fontsize=14, fontweight="bold")
    ax.set_xlabel("Revenue (₺)")
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{x:,.0f}'))
    ax.invert_yaxis()
    plt.tight_layout()
    p = os.path.join(out_dir, "top_products.png")
    fig.savefig(p, dpi=150)
    plt.close(fig)
    paths["products"] = p

    # 3 — Category Pie
    fig, ax = plt.subplots(figsize=(7, 7))
    res["category_rev"].plot(kind="pie", autopct="%1.1f%%", colors=palette, ax=ax,
                             textprops={"fontsize": 10})
    ax.set_ylabel("")
    ax.set_title("Revenue by Category", fontsize=14, fontweight="bold")
    plt.tight_layout()
    p = os.path.join(out_dir, "category_pie.png")
    fig.savefig(p, dpi=150)
    plt.close(fig)
    paths["category"] = p

    # 4 — City Revenue
    if "city_rev" in res:
        fig, ax = plt.subplots(figsize=(10, 5))
        res["city_rev"].plot(kind="bar", color=palette[3], ax=ax, edgecolor="white")
        ax.set_title("Revenue by City", fontsize=14, fontweight="bold")
        ax.set_ylabel("Revenue (₺)")
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{x:,.0f}'))
        plt.xticks(rotation=45, ha="right")
        plt.tight_layout()
        p = os.path.join(out_dir, "city_revenue.png")
        fig.savefig(p, dpi=150)
        plt.close(fig)
        paths["city"] = p

    print(f"   ✅ {len(paths)} charts saved to {out_dir}/")
    return paths


# ── STEP 4 — Excel Report ──────────────────────────────────────
def build_excel_report(df, res, chart_paths, output_path):
    print("📝 Building Excel report …")
    wb = Workbook()

    # ── Summary Sheet ───────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = ACCENT

    kpis = [
        ("KPI", "Value"),
        ("Total Revenue", f"₺{res['total_revenue']:,.2f}"),
        ("Total Orders", f"{res['total_orders']:,}"),
        ("Average Order Value", f"₺{res['avg_order']:,.2f}"),
        ("Median Order Value", f"₺{res['median_order']:,.2f}"),
        ("Report Generated", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for r, (k, v) in enumerate(kpis, 1):
        ws.cell(row=r, column=1, value=k).font = Font(bold=(r == 1))
        ws.cell(row=r, column=2, value=v).font = Font(bold=(r == 1))
        if r == 1:
            ws.cell(row=r, column=1).fill = header_fill
            ws.cell(row=r, column=1).font = header_font
            ws.cell(row=r, column=2).fill = header_fill
            ws.cell(row=r, column=2).font = header_font
    auto_width(ws)

    # Embed charts
    from openpyxl.drawing.image import Image as XlImage
    row_offset = len(kpis) + 3
    for i, (key, path) in enumerate(chart_paths.items()):
        img = XlImage(path)
        img.width = 600
        img.height = 300
        ws.add_image(img, f"A{row_offset + i * 18}")

    # ── Top Products Sheet ──────────────────────────────────────
    ws2 = wb.create_sheet("Top Products")
    ws2.sheet_properties.tabColor = GREEN
    ws2.append(["Rank", "Product", "Total Revenue (₺)"])
    style_header(ws2, 3)
    for rank, (prod, rev) in enumerate(res["top_products"].items(), 1):
        ws2.append([rank, prod, round(rev, 2)])
        ws2.cell(row=rank + 1, column=3).number_format = money_fmt
    auto_width(ws2)

    # ── Monthly Revenue Sheet ───────────────────────────────────
    if "monthly_rev" in res:
        ws3 = wb.create_sheet("Monthly Revenue")
        ws3.sheet_properties.tabColor = ORANGE
        ws3.append(["Month", "Revenue (₺)"])
        style_header(ws3, 2)
        for month, rev in res["monthly_rev"].items():
            ws3.append([str(month), round(rev, 2)])
        # Add a line chart
        chart = LineChart()
        chart.title = "Monthly Revenue Trend"
        chart.y_axis.title = "Revenue (₺)"
        chart.x_axis.title = "Month"
        chart.style = 10
        data = Reference(ws3, min_col=2, min_row=1, max_row=ws3.max_row)
        cats = Reference(ws3, min_col=1, min_row=2, max_row=ws3.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 20
        chart.height = 12
        ws3.add_chart(chart, f"D2")
        auto_width(ws3)

    # ── Raw Data Sheet ──────────────────────────────────────────
    ws4 = wb.create_sheet("Raw Data")
    # Drop Period columns that can't serialize to Excel
    df_export = df.drop(columns=["Month"], errors="ignore")
    for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
    auto_width(ws4)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"   ✅ Report saved → {output_path}")


# ── Main ────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        input_path = "sample_data/sales_data.xlsx"
        print(f"ℹ️  No input file given, using default: {input_path}")
    else:
        input_path = sys.argv[1]

    if not os.path.exists(input_path):
        print(f"❌ File not found: {input_path}")
        print("   Run `python generate_sample_data.py` first to create sample data.")
        sys.exit(1)

    output_path = "reports/sales_report.xlsx"
    charts_dir  = "reports/charts"

    df  = load_and_clean(input_path)
    res = analyse(df)
    chart_paths = create_charts(res, charts_dir)
    build_excel_report(df, res, chart_paths, output_path)

    print("\n🎉 Done! Your report is ready.")
    print(f"   📄 Excel Report : {output_path}")
    print(f"   📊 Charts       : {charts_dir}/")


if __name__ == "__main__":
    main()
