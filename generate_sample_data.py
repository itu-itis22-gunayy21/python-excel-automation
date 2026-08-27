"""
Generate realistic sample sales data for demonstration.
Creates a sample Excel file with 500+ rows of sales data.
"""

import random
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generate_sample_data(output_path="sample_data/sales_data.xlsx"):
    """Generate a realistic sales dataset."""
    
    products = {
        "Electronics": [
            ("Laptop Pro 15", 12999.99),
            ("Wireless Mouse", 299.99),
            ("USB-C Hub", 549.99),
            ("Mechanical Keyboard", 899.99),
            ("Monitor 27\"", 4999.99),
            ("Webcam HD", 699.99),
            ("External SSD 1TB", 1299.99),
            ("Headphones BT", 1499.99),
        ],
        "Office Supplies": [
            ("A4 Paper (500 sheets)", 89.99),
            ("Printer Ink Cartridge", 349.99),
            ("Desk Organizer", 149.99),
            ("Whiteboard Markers Set", 59.99),
            ("Notebook Premium", 45.99),
            ("Stapler Heavy Duty", 79.99),
        ],
        "Furniture": [
            ("Ergonomic Chair", 5499.99),
            ("Standing Desk", 7999.99),
            ("Monitor Arm", 899.99),
            ("Cable Management Kit", 199.99),
            ("Desk Lamp LED", 349.99),
        ],
        "Software": [
            ("Antivirus 1-Year", 399.99),
            ("Office Suite License", 1299.99),
            ("Cloud Storage 1TB/yr", 599.99),
            ("Project Management Tool", 249.99),
        ],
    }
    
    cities = [
        "Istanbul", "Ankara", "Izmir", "Bursa", "Antalya",
        "Adana", "Konya", "Gaziantep", "Mersin", "Kayseri"
    ]
    
    sales_reps = [
        "Ahmet Yılmaz", "Elif Kaya", "Mehmet Demir", "Zeynep Çelik",
        "Can Öztürk", "Ayşe Şahin", "Burak Arslan", "Selin Aydın",
        "Emre Doğan", "Fatma Yıldız"
    ]
    
    payment_methods = ["Credit Card", "Bank Transfer", "Cash", "Installment"]
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Data"
    
    # Headers
    headers = [
        "Order ID", "Date", "Product", "Category", "Quantity",
        "Unit Price (TL)", "Total (TL)", "City", "Sales Rep",
        "Payment Method", "Customer Type", "Discount %"
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Generate 500 rows of data
    start_date = datetime.date(2024, 1, 1)
    end_date = datetime.date(2024, 12, 31)
    delta = (end_date - start_date).days
    
    for i in range(1, 501):
        category = random.choice(list(products.keys()))
        product_name, unit_price = random.choice(products[category])
        quantity = random.randint(1, 20)
        
        # Seasonal variation
        order_date = start_date + datetime.timedelta(days=random.randint(0, delta))
        
        # Higher sales in Q4
        if order_date.month >= 10:
            quantity = int(quantity * random.uniform(1.2, 1.8))
        
        discount = random.choice([0, 0, 0, 5, 10, 15, 20])
        total = round(unit_price * quantity * (1 - discount / 100), 2)
        customer_type = random.choice(["Individual", "Corporate", "Corporate", "Reseller"])
        
        row_data = [
            f"ORD-{i:04d}",
            order_date,
            product_name,
            category,
            quantity,
            unit_price,
            total,
            random.choice(cities),
            random.choice(sales_reps),
            random.choice(payment_methods),
            customer_type,
            discount,
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=value)
            cell.border = thin_border
            if col in [6, 7]:
                cell.number_format = '#,##0.00'
            if col == 2:
                cell.number_format = 'DD/MM/YYYY'
    
    # Set column widths
    widths = [12, 12, 25, 15, 10, 15, 15, 12, 18, 15, 15, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    
    # Auto-filter
    ws.auto_filter.ref = f"A1:L501"
    
    import os
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"✅ Sample data generated: {output_path}")
    print(f"   → 500 sales records across {len(products)} categories")
    return output_path


if __name__ == "__main__":
    generate_sample_data()
